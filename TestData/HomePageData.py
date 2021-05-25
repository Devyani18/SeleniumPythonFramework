import openpyxl


class HomePageData:
    formData = [{'firstname': "Devyani", 'email': "devyaniparate1996@gmail.com", 'gender': "Female"},
                            {'firstname': "Raagu", 'email': "Raagu@lol.com", 'gender': "Male"}]

    @staticmethod
    def getTestDataFromExcel(RecordName):
        dataWorkbook = openpyxl.load_workbook("D:\\Study\\SeleniumPython\\excelDemo.xlsx")
        dataSheet = dataWorkbook.active
        PersonDetails = {}

        for i in range(1, dataSheet.max_row + 1):
            if dataSheet.cell(row=i, column=1).value == RecordName:
                for j in range(2, dataSheet.max_column):
                    PersonDetails[dataSheet.cell(row=1, column=j).value] = dataSheet.cell(row=i, column=j).value
        return PersonDetails
